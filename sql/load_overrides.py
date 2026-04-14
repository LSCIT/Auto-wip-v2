"""
load_overrides.py
Loads all 40 historical WIP override Excel files into LylesWIP.WipJobData.
Run AFTER LylesWIP_CreateDB.sql has been executed.

Usage:
    python3 sql/load_overrides.py [--dry-run]

Requirements:
    pip install openpyxl pyodbc
"""

import os
import sys
import re
import pyodbc
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation

# =============================================================================
# Config
# =============================================================================
SERVER   = '10.103.30.11'
DATABASE = 'LylesWIP'
USERNAME = 'wip.excel.sql'
PASSWORD = 'WES@2024'

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Data-from-Nicole')

COMPANY_CODES = {
    'WML':  15,   # W. M. Lyles Co.
    'AIC':  16,   # Advanced Integration & Controls
    'APC':  12,   # American Paving Co.
    'NESM': 13,   # New England Sheet Metal and Mechanical Co.
}

DRY_RUN = '--dry-run' in sys.argv

MERGE_SQL = """
MERGE dbo.WipJobData AS T
USING (VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)) AS S(
    JCCo, Job, WipMonth,
    OpsRevOverride, OpsRevPlugged, GAAPRevOverride, GAAPRevPlugged,
    OpsCostOverride, OpsCostPlugged, GAAPCostOverride, GAAPCostPlugged,
    BonusProfit, CompletionDate
)
ON T.JCCo = S.JCCo AND T.Job = S.Job AND T.WipMonth = S.WipMonth
WHEN MATCHED AND T.Source = 'ExcelImport' THEN
    UPDATE SET
        OpsRevOverride   = S.OpsRevOverride,   OpsRevPlugged   = S.OpsRevPlugged,
        GAAPRevOverride  = S.GAAPRevOverride,  GAAPRevPlugged  = S.GAAPRevPlugged,
        OpsCostOverride  = S.OpsCostOverride,  OpsCostPlugged  = S.OpsCostPlugged,
        GAAPCostOverride = S.GAAPCostOverride, GAAPCostPlugged = S.GAAPCostPlugged,
        BonusProfit      = S.BonusProfit,
        CompletionDate   = S.CompletionDate,
        UpdatedAt        = GETDATE()
WHEN NOT MATCHED THEN
    INSERT (JCCo, Job, WipMonth,
            OpsRevOverride, OpsRevPlugged, GAAPRevOverride, GAAPRevPlugged,
            OpsCostOverride, OpsCostPlugged, GAAPCostOverride, GAAPCostPlugged,
            BonusProfit, CompletionDate, UserName, Source)
    VALUES (S.JCCo, S.Job, S.WipMonth,
            S.OpsRevOverride, S.OpsRevPlugged, S.GAAPRevOverride, S.GAAPRevPlugged,
            S.OpsCostOverride, S.OpsCostPlugged, S.GAAPCostOverride, S.GAAPCostPlugged,
            S.BonusProfit, S.CompletionDate, 'ExcelImport', 'ExcelImport');
"""

# =============================================================================
# Helpers
# =============================================================================
MONTH_ABBR = {
    'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,
    'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12,
    'JANUARY':1,'FEBRUARY':2,'MARCH':3,'APRIL':4,'JUNE':6,
    'JULY':7,'AUGUST':8,'SEPTEMBER':9,'OCTOBER':10,'NOVEMBER':11,'DECEMBER':12,
}

def parse_filename(filename):
    base = os.path.splitext(filename)[0]
    jcco = None
    for prefix, code in COMPANY_CODES.items():
        if base.upper().startswith(prefix):
            jcco = code
            break
    if jcco is None:
        return None, None

    parts = base.rsplit(' - ', 1)
    if len(parts) != 2:
        return None, None
    month_str = parts[1].strip().upper()

    # Compact: DEC2025
    m = re.match(r'^([A-Z]+)(\d{4})$', month_str)
    if m:
        mon_num = MONTH_ABBR.get(m.group(1))
        if mon_num:
            return jcco, datetime(int(m.group(2)), mon_num, 1).date()

    # Long: August 2025
    m = re.match(r'^([A-Z]+)\s+(\d{4})$', month_str)
    if m:
        mon_num = MONTH_ABBR.get(m.group(1))
        if mon_num:
            return jcco, datetime(int(m.group(2)), mon_num, 1).date()

    return None, None

def to_decimal(value):
    """Returns float if value is numeric, else None (NULL = no override).
    Explicit zeros are preserved — a $0 override is meaningful in accounting."""
    if value is None:
        return None
    try:
        d = Decimal(str(value))
        return float(d)
    except (InvalidOperation, TypeError):
        return None

def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    return None

def normalize_job(raw):
    """Ensure Vista-format: company.NNNN. where NNNN is 4-digit job number.
    Pads short job segments: '73.105' → '73.1050.' (right-pad with zeros to 4 digits)."""
    s = str(raw).strip()
    if not s.endswith('.'):
        s = s + '.'
    # Pad job number segment to 4 digits: '73.105.' → '73.1050.'
    parts = s.rstrip('.').split('.', 1)
    if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) < 4:
        s = parts[0] + '.' + parts[1].ljust(4, '0') + '.'
    return s

# =============================================================================
# Read one file → dict keyed by job number
# =============================================================================
def read_file(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    rows = {}

    if 'Revenue' in wb.sheetnames:
        for row in wb['Revenue'].iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            job = normalize_job(row[1])
            rows.setdefault(job, {})
            rows[job].update({
                'gaap_rev':   to_decimal(row[3]),
                'ops_rev':    to_decimal(row[4]),
                'bonus':      to_decimal(row[5]),
                'completion': to_date(row[6]),
            })

    if 'Cost' in wb.sheetnames:
        for row in wb['Cost'].iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            job = normalize_job(row[1])
            rows.setdefault(job, {})
            rows[job].update({
                'gaap_cost': to_decimal(row[3]),
                'ops_cost':  to_decimal(row[4]),
            })

    wb.close()
    return rows

# =============================================================================
# Main
# =============================================================================
def main():
    files = sorted([
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ])

    print(f"Found {len(files)} files")
    if DRY_RUN:
        print("DRY RUN — no database writes\n")

    conn = None if DRY_RUN else pyodbc.connect(
        f"DRIVER={{ODBC Driver 18 for SQL Server}};"
        f"SERVER={SERVER};DATABASE={DATABASE};"
        f"UID={USERNAME};PWD={PASSWORD};TrustServerCertificate=yes;Encrypt=no;"
    )

    total_rows = 0
    skipped = 0
    errors = []

    for filename in files:
        jcco, wip_month = parse_filename(filename)
        if jcco is None:
            print(f"  SKIP: {filename}")
            skipped += 1
            continue

        company_label = next(k for k, v in COMPANY_CODES.items() if v == jcco)
        filepath = os.path.join(DATA_DIR, filename)

        try:
            job_data = read_file(filepath)
        except Exception as e:
            errors.append(f"{filename}: {e}")
            print(f"  ERROR reading {filename}: {e}")
            continue

        print(f"  {company_label} Co={jcco} {wip_month:%Y-%m}: {len(job_data)} jobs")

        if DRY_RUN:
            total_rows += len(job_data)
            continue

        params = [
            (
                jcco, job, wip_month,
                d.get('ops_rev'),   1 if d.get('ops_rev')   is not None else 0,
                d.get('gaap_rev'),  1 if d.get('gaap_rev')  is not None else 0,
                d.get('ops_cost'),  1 if d.get('ops_cost')  is not None else 0,
                d.get('gaap_cost'), 1 if d.get('gaap_cost') is not None else 0,
                d.get('bonus'),
                d.get('completion'),
            )
            for job, d in job_data.items()
        ]

        cursor = conn.cursor()
        try:
            cursor.executemany(MERGE_SQL, params)
            conn.commit()
            total_rows += len(params)
        except Exception as e:
            conn.rollback()
            errors.append(f"{filename}: {e}")
            print(f"    DB ERROR: {e}")

    if conn:
        conn.close()

    print(f"\nDone. {total_rows} rows upserted across {len(files) - skipped - len(errors)} files.")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors:
            print(f"  {e}")

if __name__ == '__main__':
    main()
