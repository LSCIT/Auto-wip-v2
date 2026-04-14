"""
validate_wip.py
Compares Nicole's Dec 2025 WIP import files against workbook snapshots.
Outputs Excel report for Nicole/Cindy review.

Usage:
    python3 validate_wip.py

Output:
    vm/validation_report.xlsx
"""

import os
import sys
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Windows console defaults to cp1252 — force UTF-8 so Unicode chars print correctly
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(BASE_DIR, 'Data-from-Nicole')
# On Windows VM (WSL): BatchValidate saves to C:\Trusted\validate-d3\ = /mnt/c/Trusted/validate-d3/
# On Mac: snapshots were copied to vm/validate-d3/ relative to project
SNAPSHOT_DIR = os.environ.get('WIP_SNAPSHOT_DIR',
               '/mnt/c/Trusted/validate-d3' if os.path.isdir('/mnt/c/Trusted/validate-d3')
               else os.path.join(BASE_DIR, 'vm', 'validate-d3'))
OUTPUT_FILE  = os.path.join(BASE_DIR, 'vm', 'validation_report.xlsx')

TOLERANCE = 0.02   # $0.02 tolerance for float precision

COMPANY_NAMES = {15: 'WML', 16: 'AIC', 12: 'APC', 13: 'NESM'}

DEC_FILES = {
    15: 'WML WIP History Import - DEC2025.xlsx',
    16: 'AIC WIP History Import - December 2025.xlsx',
    12: 'APC WIP History Import - December 2025.xlsx',
    13: 'NESM WIP History Import - December 2025.xlsx',
}

DIVISIONS = {
    15: [50, 51, 52, 53, 54, 55, 56, 57, 58],   # Div50 added: Nicole has 7 jobs (50.0541–50.0551)
    16: [70, 71, 72, 73, 74, 75, 76, 77, 78],
    12: [20, 21],                                  # Div20 added: Nicole has 10 jobs (20.252–20.2574)
    13: [31, 32, 33, 35],
}

# (label, nicole_key, ops_z_col, gaap_z_col)
CHECKS = [
    ('Ops Rev',   'ops_rev',   'COLZOPsRev',  None),
    ('Ops Cost',  'ops_cost',  'COLZOPsCost', None),
    ('GAAP Rev',  'gaap_rev',  None,          'COLZGAAPRev'),
    ('GAAP Cost', 'gaap_cost', None,          'COLZGAAPCost'),
    ('Bonus',     'bonus',     'COLZOPsBonus',None),
]

# ── Colors ────────────────────────────────────────────────────────────────────
GREEN  = PatternFill('solid', fgColor='C6EFCE')
RED    = PatternFill('solid', fgColor='FFC7CE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')
GRAY   = PatternFill('solid', fgColor='EDEDED')
HDR_FILL = PatternFill('solid', fgColor='1F3864')

GREEN_FONT  = Font(color='276221')
RED_FONT    = Font(color='9C0006')
YELLOW_FONT = Font(color='9C6500')
BOLD        = Font(bold=True)
HDR_FONT    = Font(bold=True, color='FFFFFF')

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_job(raw):
    s = str(raw).strip()
    if not s.endswith('.'):
        s += '.'
    parts = s.rstrip('.').split('.', 1)
    if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) < 4:
        s = parts[0] + '.' + parts[1].ljust(4, '0') + '.'
    return s

def to_float(val):
    if val is None:
        return None
    try:
        return float(Decimal(str(val)))
    except (InvalidOperation, TypeError, ValueError):
        return None

def to_date_val(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip().upper()
        if s in ('TBD', '', 'NONE', 'N/A'):
            return None
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                pass
    return None

def fmt_num(v):
    f = to_float(v)
    if f is None:
        return ''
    return f'{f:,.2f}'

def fmt_date(v):
    d = to_date_val(v) if not isinstance(v, date) else v
    return str(d)[:10] if d else ''

def has_override(nicole_job_dict):
    """True if Nicole provided at least one non-zero override for this job."""
    for _, nkey, _, _ in CHECKS:
        v = to_float(nicole_job_dict.get(nkey))
        if v is not None and abs(v) >= TOLERANCE:
            return True
    if nicole_job_dict.get('completion') is not None:
        return True
    return False

# ── Read Nicole file ──────────────────────────────────────────────────────────

def read_nicole(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    jobs = {}
    if 'Revenue' in wb.sheetnames:
        for row in wb['Revenue'].iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            job = normalize_job(row[1])
            jobs.setdefault(job, {})
            jobs[job].update({
                'gaap_rev':   to_float(row[3]),
                'ops_rev':    to_float(row[4]),
                'bonus':      to_float(row[5]),
                'completion': to_date_val(row[6]),
            })
    if 'Cost' in wb.sheetnames:
        for row in wb['Cost'].iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            job = normalize_job(row[1])
            jobs.setdefault(job, {})
            jobs[job].update({
                'gaap_cost': to_float(row[3]),
                'ops_cost':  to_float(row[4]),
            })
    wb.close()
    return jobs

# ── Read snapshot sheet ───────────────────────────────────────────────────────

def read_snapshot_sheet(filepath, sheet_name):
    if not os.path.exists(filepath):
        return {}, {}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}, {}
    ws = wb[sheet_name]
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = {v: i for i, v in enumerate(row1) if v}
    if 'COLJobNumber' not in col_map:
        wb.close()
        return col_map, {}
    jn_idx = col_map['COLJobNumber']
    jobs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[jn_idx] if len(row) > jn_idx else None
        if raw is None:
            continue
        s = str(raw).strip()
        if '.' not in s or s.startswith('Job'):
            continue
        if not s.replace('.', '').replace(' ', '').isdigit():
            continue
        if not s.endswith('.'):
            s += '.'
        jobs[s] = row
    wb.close()
    return col_map, jobs

def get_cell(row, col_map, col_name):
    if col_name not in col_map or row is None:
        return None
    idx = col_map[col_name]
    return row[idx] if idx < len(row) else None

# ── Compare one job ───────────────────────────────────────────────────────────

def check_match(nicole_val, wb_z_val):
    n = to_float(nicole_val)
    if n is None or abs(n) < TOLERANCE:
        return 'NO_OVERRIDE'
    w = to_float(wb_z_val)
    if w is None:
        return 'MISMATCH'
    return 'MATCH' if abs(n - w) <= TOLERANCE else 'MISMATCH'

def check_date_match(nicole_date, wb_val):
    if nicole_date is None:
        return 'NO_OVERRIDE'
    wb_d = to_date_val(wb_val)
    if wb_d is None:
        return 'MISMATCH'
    return 'MATCH' if nicole_date == wb_d else 'MISMATCH'

def compare_job(job, n_dict, ops_row, ops_cm, gaap_row, gaap_cm):
    """Returns dict of {label: (nicole_val, wb_val, status)} for all checks."""
    results = {}
    any_mismatch = False
    any_override = False

    for label, nkey, ops_col, gaap_col in CHECKS:
        n_val = n_dict.get(nkey)
        if ops_col:
            wb_val = get_cell(ops_row, ops_cm, ops_col)
        else:
            wb_val = get_cell(gaap_row, gaap_cm, gaap_col)
        status = check_match(n_val, wb_val)
        results[label] = (n_val, wb_val, status)
        if status != 'NO_OVERRIDE':
            any_override = True
        if status == 'MISMATCH':
            any_mismatch = True

    # Completion date (from Jobs-Ops)
    n_date   = n_dict.get('completion')
    wb_d_val = get_cell(ops_row, ops_cm, 'COLCompDate')
    d_status = check_date_match(n_date, wb_d_val)
    results['Completion'] = (n_date, wb_d_val, d_status)
    if d_status != 'NO_OVERRIDE':
        any_override = True
    if d_status == 'MISMATCH':
        any_mismatch = True

    return results, any_override, any_mismatch

# ── Validate one division ─────────────────────────────────────────────────────

def validate_division(co, div, nicole_jobs, ops_cm, ops_rows, gaap_cm, gaap_rows):
    """
    Compares WB snapshot jobs against Nicole's data.
    'Missing' is NOT reported here — it's tracked company-wide.
    Returns list of per-job result dicts + division summary.
    """
    results = []
    wb_job_set = set(ops_rows.keys()) | set(gaap_rows.keys())

    matched   = 0
    mismatched= 0
    no_ovr    = 0   # WB jobs with no Nicole override (Vista-only)
    not_found = 0   # WB jobs where Nicole job key not in Nicole file at all

    for job in sorted(wb_job_set):
        ops_row  = ops_rows.get(job)
        gaap_row = gaap_rows.get(job)
        n_dict   = nicole_jobs.get(job, {})
        in_nicole = job in nicole_jobs

        if not in_nicole:
            not_found += 1
            # Still record so report can show Vista-only jobs
            results.append({
                'co': co, 'div': div, 'job': job,
                'in_nicole': False,
                'any_override': False,
                'any_mismatch': False,
                'checks': {},
                'vista_only': True,
            })
            continue

        checks, any_ovr, any_mm = compare_job(job, n_dict, ops_row, ops_cm, gaap_row, gaap_cm)

        if any_mm:
            mismatched += 1
        elif any_ovr:
            matched += 1
        else:
            no_ovr += 1

        results.append({
            'co': co, 'div': div, 'job': job,
            'in_nicole': True,
            'any_override': any_ovr,
            'any_mismatch': any_mm,
            'checks': checks,
        })

    summary = {
        'co': co, 'div': div,
        'wb_jobs': len(wb_job_set),
        'matched': matched,
        'mismatched': mismatched,
        'no_nicole_override': no_ovr,
        'vista_only': not_found,
        'nicole_jobs_seen': wb_job_set & set(nicole_jobs.keys()),
    }
    return results, summary

# ── Excel report ──────────────────────────────────────────────────────────────

DETAIL_HEADERS = [
    'Company', 'Division', 'Job',
    'Nicole Ops Rev', 'WB Z Ops Rev', 'Ops Rev',
    'Nicole Ops Cost', 'WB Z Ops Cost', 'Ops Cost',
    'Nicole GAAP Rev', 'WB Z GAAP Rev', 'GAAP Rev',
    'Nicole GAAP Cost', 'WB Z GAAP Cost', 'GAAP Cost',
    'Nicole Bonus', 'WB Z Bonus', 'Bonus',
    'Nicole Completion', 'WB Completion', 'Date',
    'Note',
]
STATUS_COLS = {6, 9, 12, 15, 18, 21}   # 1-indexed columns that hold status text

def add_header_row(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[row].height = 28

def color_status(cell, status):
    if status == 'MATCH':
        cell.fill = GREEN; cell.font = GREEN_FONT
    elif status == 'MISMATCH':
        cell.fill = RED;   cell.font = RED_FONT
    elif status == 'NO_OVERRIDE':
        cell.fill = GRAY

def write_detail_row(ws, res, note=''):
    c = res['checks']

    def cv(label):
        return c.get(label, (None, None, 'NO_OVERRIDE'))

    ops_rev  = cv('Ops Rev')
    ops_cost = cv('Ops Cost')
    gaap_rev = cv('GAAP Rev')
    gaap_cost= cv('GAAP Cost')
    bonus    = cv('Bonus')
    comp     = cv('Completion')

    def wb_num(v):
        return fmt_num(v) if v is not None else ''

    row_vals = [
        COMPANY_NAMES.get(res['co'], res['co']),
        res['div'],
        res['job'],
        fmt_num(ops_rev[0]),   wb_num(ops_rev[1]),   ops_rev[2],
        fmt_num(ops_cost[0]),  wb_num(ops_cost[1]),  ops_cost[2],
        fmt_num(gaap_rev[0]),  wb_num(gaap_rev[1]),  gaap_rev[2],
        fmt_num(gaap_cost[0]), wb_num(gaap_cost[1]), gaap_cost[2],
        fmt_num(bonus[0]),     wb_num(bonus[1]),     bonus[2],
        fmt_date(comp[0]),     fmt_date(comp[1]),    comp[2],
        note,
    ]

    r = ws.max_row + 1
    for col_idx, v in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=col_idx, value=v)
        if col_idx in STATUS_COLS:
            color_status(cell, v)

def auto_width(ws, min_w=8, max_w=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

def build_report(all_summaries, all_results, company_missing):
    wb = openpyxl.Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    sum_hdrs = [
        'Company', 'Division', 'WB Jobs',
        'Override Match', 'Override Mismatch',
        'No Override (Vista only in WB)', 'Vista-only (not in Nicole)',
        'Status'
    ]
    add_header_row(ws_sum, sum_hdrs)

    for s in all_summaries:
        if 'skipped' in s:
            row_vals = [COMPANY_NAMES.get(s['co'], s['co']), s['div'],
                        '—', '—', '—', '—', '—', 'SNAPSHOT MISSING']
        else:
            status = 'PASS ✓' if s['mismatched'] == 0 else f"FAIL ✗ ({s['mismatched']})"
            row_vals = [
                COMPANY_NAMES.get(s['co'], s['co']),
                s['div'],
                s['wb_jobs'],
                s['matched'],
                s['mismatched'],
                s['no_nicole_override'],
                s['vista_only'],
                status,
            ]
        r = ws_sum.max_row + 1
        for c_idx, v in enumerate(row_vals, 1):
            ws_sum.cell(row=r, column=c_idx, value=v)
        status_cell = ws_sum.cell(row=r, column=8)
        sv = str(status_cell.value)
        if 'PASS' in sv:
            status_cell.fill = GREEN; status_cell.font = GREEN_FONT
        elif 'FAIL' in sv or 'MISSING' in sv:
            status_cell.fill = RED;   status_cell.font = RED_FONT

    # Totals
    numeric_sums = {k: 0 for k in ['wb_jobs','matched','mismatched','no_nicole_override','vista_only']}
    for s in all_summaries:
        if 'skipped' not in s:
            for k in numeric_sums:
                numeric_sums[k] += s[k]
    r = ws_sum.max_row + 1
    ws_sum.cell(row=r, column=1, value='TOTAL').font = BOLD
    for c_idx, k in enumerate(['wb_jobs','matched','mismatched','no_nicole_override','vista_only'], 3):
        ws_sum.cell(row=r, column=c_idx, value=numeric_sums[k]).font = BOLD

    # Company-level missing
    r += 2
    ws_sum.cell(row=r, column=1, value='Jobs in Nicole\'s file not found in ANY division snapshot:').font = BOLD
    r += 1
    for co, missing_jobs in sorted(company_missing.items()):
        label = COMPANY_NAMES.get(co, co)
        ws_sum.cell(row=r, column=1, value=label).font = BOLD
        ws_sum.cell(row=r, column=2, value=f'{len(missing_jobs)} jobs not in any WB snapshot')
        if missing_jobs:
            ws_sum.cell(row=r, column=3,
                        value=', '.join(sorted(missing_jobs)[:20]) + (' ...' if len(missing_jobs) > 20 else ''))
        r += 1

    auto_width(ws_sum)

    # ── Mismatches ────────────────────────────────────────────────────────────
    ws_mm = wb.create_sheet('Mismatches')
    add_header_row(ws_mm, DETAIL_HEADERS)

    mismatch_results = [r for r in all_results if r.get('any_mismatch')]
    if mismatch_results:
        for res in mismatch_results:
            write_detail_row(ws_mm, res)
    else:
        ws_mm.cell(row=2, column=1,
                   value='No mismatches found — all overrides match ✓').font = Font(bold=True, color='276221')

    auto_width(ws_mm)

    # ── All Override Checks ───────────────────────────────────────────────────
    ws_all = wb.create_sheet('All Override Checks')
    add_header_row(ws_all, DETAIL_HEADERS)

    for res in all_results:
        if not res.get('in_nicole') or not res.get('any_override'):
            continue
        write_detail_row(ws_all, res)

    auto_width(ws_all)

    # ── Company-level missing jobs ────────────────────────────────────────────
    ws_miss = wb.create_sheet('Nicole Jobs Not In WB')
    miss_hdrs = ['Company', 'Job', 'Nicole Ops Rev', 'Nicole Ops Cost',
                 'Nicole GAAP Rev', 'Nicole GAAP Cost', 'Nicole Bonus', 'Nicole Completion']
    add_header_row(ws_miss, miss_hdrs)

    for co, missing_jobs in sorted(company_missing.items()):
        label = COMPANY_NAMES.get(co, co)
        # We need the Nicole data — re-read or pass it in
        for job in sorted(missing_jobs):
            ws_miss.cell(row=ws_miss.max_row + 1, column=1, value=label)
            ws_miss.cell(row=ws_miss.max_row, column=2, value=job)

    auto_width(ws_miss)

    # ── Notes ─────────────────────────────────────────────────────────────────
    ws_n = wb.create_sheet('Notes')
    notes = [
        ('Scope', 'Dec 2025 only. Compares Nicole WIP History Import files vs workbook snapshots loaded from Vista + LylesWIP.'),
        ('Ground truth', 'Z columns (COLZOPsRev, COLZGAAPRev, etc.) store exactly what LylesWIP wrote to the workbook. If they match Nicole\'s source values, the pipeline is correct end-to-end.'),
        ('NO_OVERRIDE', 'Nicole put 0 or blank → no override in LylesWIP → workbook shows Vista-calculated value. Not a problem.'),
        ('MATCH', 'Nicole\'s non-zero value matches the workbook Z column within $0.02.'),
        ('MISMATCH', 'Nicole had a non-zero value but workbook Z column differs or is empty. Needs investigation.'),
        ('Vista-only jobs', 'Jobs the workbook shows from Vista that Nicole did not include in her file. No override expected — Vista values are used. Normal behavior.'),
        ('Nicole Jobs Not In WB', 'Jobs Nicole provided that did not appear in any division snapshot. May mean: job is closed in Vista, job number mismatch, or the division snapshot is missing.'),
        ('Jobs-Ops vs GAAP', 'Not validated here — that sheet is not loaded by the batch process. Validate manually in the live workbook.'),
        ('Tolerance', f'$0.02 — values within this range are considered a match.'),
    ]
    ws_n.cell(row=1, column=1, value='Term').font = BOLD
    ws_n.cell(row=1, column=2, value='Explanation').font = BOLD
    for r_idx, (k, v) in enumerate(notes, 2):
        ws_n.cell(row=r_idx, column=1, value=k).font = BOLD
        ws_n.cell(row=r_idx, column=2, value=v).alignment = Alignment(wrap_text=True)
    ws_n.column_dimensions['A'].width = 22
    ws_n.column_dimensions['B'].width = 85

    wb.save(OUTPUT_FILE)
    print(f'\nReport saved: {OUTPUT_FILE}')

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    all_summaries  = []
    all_results    = []
    company_missing= {}   # co → set of job numbers in Nicole but found in NO division snapshot

    for co in [15, 16, 12, 13]:
        label = COMPANY_NAMES[co]
        nicole_file = os.path.join(DATA_DIR, DEC_FILES[co])
        if not os.path.exists(nicole_file):
            print(f'SKIP {label}: Nicole file not found')
            continue

        nicole_jobs = read_nicole(nicole_file)
        print(f'\n{label} (Co{co}): {len(nicole_jobs)} jobs in Nicole Dec 2025')

        # Track Nicole jobs found across all divisions
        nicole_seen = set()

        for div in DIVISIONS[co]:
            snap = os.path.join(SNAPSHOT_DIR, f'{co}-{div}.xltm')
            if not os.path.exists(snap):
                print(f'  Div{div}: snapshot not found — SKIP')
                all_summaries.append({'co': co, 'div': div, 'skipped': True})
                continue

            ops_cm,  ops_rows  = read_snapshot_sheet(snap, 'Jobs-Ops')
            gaap_cm, gaap_rows = read_snapshot_sheet(snap, 'Jobs-GAAP')

            results, summary = validate_division(
                co, div, nicole_jobs,
                ops_cm, ops_rows, gaap_cm, gaap_rows
            )
            all_summaries.append(summary)
            all_results.extend(results)

            nicole_seen.update(summary['nicole_jobs_seen'])

            mm = summary['mismatched']
            status = 'PASS' if mm == 0 else f'FAIL ({mm} mismatches)'
            print(f'  Div{div}: {summary["wb_jobs"]} WB jobs | '
                  f'{summary["matched"]} matched | '
                  f'{mm} mismatched | '
                  f'{summary["vista_only"]} Vista-only — {status}')

        # Company-level: Nicole jobs not seen in any division
        nicole_with_overrides = {j for j, d in nicole_jobs.items() if has_override(d)}
        truly_missing = nicole_with_overrides - nicole_seen
        company_missing[co] = truly_missing

        if truly_missing:
            print(f'  → {len(truly_missing)} Nicole jobs w/ overrides not found in any division snapshot')
        else:
            print(f'  → All Nicole jobs w/ overrides found in at least one snapshot ✓')

    # Populate missing jobs detail sheet with Nicole values
    # (we need to re-read Nicole files for the detail sheet)
    for co in company_missing:
        if not company_missing[co]:
            continue
        nicole_file = os.path.join(DATA_DIR, DEC_FILES[co])
        nicole_jobs = read_nicole(nicole_file)
        # We'll pass this into the report builder via company_missing_detail
        company_missing[co] = {j: nicole_jobs[j] for j in company_missing[co] if j in nicole_jobs}

    build_report(all_summaries, all_results, company_missing)

    total_mm  = sum(s.get('mismatched', 0) for s in all_summaries)
    total_div = len(all_summaries)
    passing   = sum(1 for s in all_summaries if s.get('mismatched', 1) == 0 and 'skipped' not in s)
    print(f'\n{passing}/{total_div} divisions pass. Total mismatches: {total_mm}')

if __name__ == '__main__':
    main()
